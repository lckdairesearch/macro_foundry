"""Generate docs/schema/category_taxonomy.xlsx from the reviewed/synthesized
macrodb category taxonomy (ADR 0026 + workflow review w0ul9g3el).

This is a planning/seed-scaffold artifact, not application code. Re-run to
regenerate the workbook after editing the TAXONOMY data below.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# (l1, l2, kind, [l3...], note)  -- note carries provenance for changed rows.
# Series-flag collapses applied (the dimension noted moves to a series-level flag).
TAXONOMY = [
    ("PRICES", "CONSUMER_PRICES", "topic",
     ["CPI_ALL_ITEMS", "CPI_CORE"],
     "Collapsed CPI_FOOD/ENERGY/SERVICES -> series flag: component (COICOP). CPI_CORE kept per ADR 0026 §5.2."),
    ("PRICES", "PRODUCER_PRICES", "topic", ["PPI_ALL_ITEMS", "PPI_INPUT", "PPI_OUTPUT"],
     "Input/output kept (stage-of-processing = genuine distinction)."),
    ("PRICES", "TRADE_PRICES", "topic", ["TRADE_PRICE_INDEX", "TERMS_OF_TRADE"],
     "Collapsed IMPORT/EXPORT_PRICE_INDEX -> TRADE_PRICE_INDEX + series flag: direction."),
    ("PRICES", "PROPERTY_PRICES", "topic",
     ["PROPERTY_PRICE_INDEX", "RENT_INDEX"],
     "Collapsed RESIDENTIAL/COMMERCIAL -> PROPERTY_PRICE_INDEX + series flag: property type. Rent kept (capital value vs rent = distinct)."),
    ("PRICES", "COMMODITY_PRICE", "concept", [],
     "Singularized concept-leaf code (final review). Series flags: commodity (oil/gas/gold/metals/agri) + measure (spot/index)."),
    ("PRICES", "COMPARATIVE_PRICE_LEVELS", "topic",
     ["PRICE_LEVEL_INDEX", "PURCHASING_POWER_PARITY"],
     "Spatial/PPP price-level comparison, distinct from temporal CPI/PPI (validation wz3k53mcb)."),

    ("NATIONAL_ACCOUNTS", "GDP_AND_GROWTH", "topic",
     ["GDP_NOMINAL", "GDP_REAL", "GDP_DEFLATOR"],
     "Collapsed GDP_PER_CAPITA -> series flag: per-capita normalization (real/nominal kept by ADR convention). ROUTING: value-added by industry -> SECTORAL_OUTPUT; regional -> GDP_* with region series dimension."),
    ("NATIONAL_ACCOUNTS", "EXPENDITURE_COMPONENTS", "topic",
     ["HOUSEHOLD_CONSUMPTION", "GOVERNMENT_CONSUMPTION", "GROSS_FIXED_CAPITAL_FORMATION", "CHANGE_IN_INVENTORIES", "NET_EXPORTS"],
     "Kept: genuine SNA components. NET_EXPORTS vs EXPORTS/IMPORTS split = open question."),
    ("NATIONAL_ACCOUNTS", "INCOME_AND_SAVING", "topic",
     ["GROSS_NATIONAL_INCOME", "OPERATING_SURPLUS", "NATIONAL_SAVING", "HOUSEHOLD_SAVING"],
     "SAVING_RATE -> SAVING (#4). OPERATING_SURPLUS moved here as income-approach GDP component (final review)."),
    ("NATIONAL_ACCOUNTS", "SECTOR_ACCOUNTS", "topic",
     ["HOUSEHOLD_SECTOR_ACCOUNT", "CORPORATE_SECTOR_ACCOUNT", "FLOW_OF_FUNDS"], ""),
    ("NATIONAL_ACCOUNTS", "INPUT_OUTPUT", "concept", [],
     "Concept-leaf; kept (decided #8)."),
    ("NATIONAL_ACCOUNTS", "NATIONAL_WEALTH", "topic",
     ["CAPITAL_STOCK", "NATIONAL_BALANCE_SHEET", "HOUSEHOLD_NET_WORTH"],
     "HOUSEHOLD_NET_WORTH kept as a watched-headline exception (Fed Z.1-style); sector is otherwise a series flag (final review)."),

    ("PRODUCTION_BUSINESS_ACTIVITY", "INDUSTRIAL_PRODUCTION", "topic",
     ["INDUSTRIAL_PRODUCTION_INDEX", "CAPACITY_UTILIZATION", "INVENTORIES"],
     "Collapsed MANUFACTURING_PRODUCTION -> sector flag. Added INVENTORIES (stock; sector + inventory-to-sales = series flags) (#6)."),
    ("PRODUCTION_BUSINESS_ACTIVITY", "SECTORAL_OUTPUT", "concept", [],
     "Collapsed to one concept -> series flag: sector (agri/mining/manufacturing/services). Receives value-added by industry (GDP_REAL routing)."),
    ("PRODUCTION_BUSINESS_ACTIVITY", "CONSTRUCTION", "topic",
     ["CONSTRUCTION_OUTPUT", "HOUSING_START", "BUILDING_PERMIT"],
     "Kept: distinct units. Singularized codes (#1)."),
    ("PRODUCTION_BUSINESS_ACTIVITY", "BUSINESS_SURVEYS", "topic",
     ["PMI", "BUSINESS_CONFIDENCE", "NEW_ORDERS"],
     "Collapsed sector PMIs -> PMI + series flag: sector. Collapsed EXPORT_ORDERS -> series flag: scope on NEW_ORDERS."),
    ("PRODUCTION_BUSINESS_ACTIVITY", "BUSINESS_DEMOGRAPHY", "topic",
     ["BUSINESS_REGISTRATION", "BUSINESS_BANKRUPTCY", "ENTERPRISE_COUNT"],
     "Singularized codes (#1)."),
    ("PRODUCTION_BUSINESS_ACTIVITY", "BUSINESS_PROFITS", "topic",
     ["CORPORATE_PROFIT"],
     "Added (validation wz3k53mcb): strongest gap (ABS + China NBS). Firm-level profits; OPERATING_SURPLUS moved to NATIONAL_ACCOUNTS (final review)."),

    ("RETAIL_CONSUMPTION", "DISTRIBUTIVE_TRADE", "topic",
     ["RETAIL_SALES", "WHOLESALE_TRADE"],
     "Retail + wholesale (distributive trade). MOTOR_VEHICLE_SALES collapsed -> product-category flag on RETAIL_SALES (final-review follow-up). RETAIL_SALES value/volume = basis flag; L2 named DISTRIBUTIVE_TRADE to avoid collision with child RETAIL_SALES."),
    ("RETAIL_CONSUMPTION", "CONSUMER_CONFIDENCE", "concept", [], ""),
    ("RETAIL_CONSUMPTION", "HOUSEHOLD_SPENDING", "concept", [],
     "Household-budget-survey spending (vs NA HOUSEHOLD_CONSUMPTION = SNA final-consumption aggregate) (#3). Category = series flag."),

    ("LABOR", "EMPLOYMENT_AND_UNEMPLOYMENT", "topic",
     ["UNEMPLOYMENT_RATE", "EMPLOYMENT_LEVEL", "PARTICIPATION_RATE", "JOB_VACANCIES", "HOURS_WORKED"],
     "Labor-market quantities. Nested JOB_VACANCIES (demand) + HOURS_WORKED (input) here as L3 for comparable L2 breadth (final-review follow-up). Collapsed YOUTH_UNEMPLOYMENT_RATE -> age/group flag; public/by-sector -> sector flag on EMPLOYMENT_LEVEL."),
    ("LABOR", "WAGES_AND_EARNINGS", "topic",
     ["AVERAGE_EARNINGS", "MINIMUM_WAGE", "LABOR_COST_INDEX"], ""),
    ("LABOR", "PRODUCTIVITY", "topic",
     ["LABOR_PRODUCTIVITY", "UNIT_LABOR_COST"], ""),

    ("MONETARY_BANKING", "MONETARY_AGGREGATES", "topic",
     ["MONETARY_BASE", "NARROW_MONEY", "BROAD_MONEY"],
     "Kept: canonical money tiers (M-letters are the series flags)."),
    ("MONETARY_BANKING", "INTEREST_RATES", "topic",
     ["POLICY_RATE", "INTERBANK_RATE", "LENDING_RATE", "DEPOSIT_RATE"],
     "Rates by type: policy (central-bank set), interbank (money market), lending/deposit (bank retail). Market-determined bond yields stay in FINANCIAL_INDICATORS (final-review follow-up)."),
    ("MONETARY_BANKING", "CREDIT_AND_DEBT", "topic",
     ["CREDIT_TO_PRIVATE_SECTOR"],
     "Collapsed HOUSEHOLD/CORPORATE/MORTGAGE_DEBT -> series flag: borrower sector + instrument."),
    ("MONETARY_BANKING", "CENTRAL_BANK_BALANCE_SHEET", "topic",
     ["CENTRAL_BANK_ASSET", "BANK_RESERVE"], "Singularized codes (#1)."),
    ("MONETARY_BANKING", "BANKING_SECTOR", "topic",
     ["BANK_DEPOSIT", "NON_PERFORMING_LOAN", "BANK_CAPITAL_ADEQUACY_RATIO", "BANK_PROFITABILITY", "BANK_LIQUIDITY_RATIO"],
     "Bank balance-sheet & soundness (IMF FSI). Rates moved to INTEREST_RATES (final-review follow-up). ROA/ROE -> BANK_PROFITABILITY + measure flag."),
    ("MONETARY_BANKING", "FINANCIAL_INCLUSION", "topic",
     ["FINANCIAL_ACCESS_POINT", "ACCOUNT_OWNERSHIP", "MOBILE_MONEY_ACCOUNT"],
     "Folded ATM_DENSITY + BANK_BRANCH_DENSITY -> FINANCIAL_ACCESS_POINT (type ATM/branch + per-capita/per-area = series flags); removes ATM abbrev + density normalization (final review)."),

    ("FINANCIAL_INDICATORS", "EQUITY_MARKETS", "topic",
     ["EQUITY_MARKET_INDEX", "MARKET_CAPITALIZATION", "DIVIDEND_YIELD"], ""),
    ("FINANCIAL_INDICATORS", "BOND_YIELDS_AND_SPREADS", "topic",
     ["GOVERNMENT_BOND_YIELD", "CORPORATE_BOND_YIELD", "SOVEREIGN_YIELD_SPREAD"],
     "Maturity already a series flag. Issuer (govt/corporate) kept (distinct asset classes)."),
    ("FINANCIAL_INDICATORS", "VOLATILITY_AND_RISK", "topic",
     ["EQUITY_VOLATILITY_INDEX", "SOVEREIGN_CDS_SPREAD"], ""),
    ("FINANCIAL_INDICATORS", "FINANCIAL_CONDITIONS", "concept", [], ""),

    ("GOVERNMENT_FISCAL", "REVENUE_AND_TAXATION", "topic",
     ["GOVERNMENT_REVENUE", "TAX_REVENUE"],
     "Kept TAX_REVENUE (headline in its own right); tax-by-type = series flag."),
    ("GOVERNMENT_FISCAL", "EXPENDITURE", "topic",
     ["GOVERNMENT_EXPENDITURE", "INTEREST_PAYMENT"],
     "Kept INTEREST_PAYMENT (debt service watched standalone); other COFOG functions = series flag. Singularized (#1)."),
    ("GOVERNMENT_FISCAL", "FISCAL_BALANCE", "concept", [], ""),
    ("GOVERNMENT_FISCAL", "PUBLIC_DEBT", "topic",
     ["GENERAL_GOVERNMENT_DEBT"], ""),

    ("INTERNATIONAL", "MERCHANDISE_TRADE", "topic",
     ["MERCHANDISE_TRADE_FLOW", "TRADE_BALANCE_GOODS"],
     "Collapsed EXPORTS/IMPORTS_GOODS -> MERCHANDISE_TRADE_FLOW + series flag: direction."),
    ("INTERNATIONAL", "TRADE_IN_SERVICES", "topic",
     ["SERVICES_TRADE_FLOW", "SERVICES_BALANCE"],
     "Collapsed EXPORTS/IMPORTS_SERVICES -> SERVICES_TRADE_FLOW + series flag: direction."),
    ("INTERNATIONAL", "BALANCE_OF_PAYMENTS", "topic",
     ["CURRENT_ACCOUNT_BALANCE", "FINANCIAL_ACCOUNT_BALANCE", "REMITTANCES"],
     "Kept REMITTANCES (EM headline)."),
    ("INTERNATIONAL", "EXCHANGE_RATE", "concept", [],
     "Singularized concept-leaf code (final review). Series flags: basis (nominal/real) + scope (bilateral/effective)."),
    ("INTERNATIONAL", "RESERVES_AND_INVESTMENT_POSITION", "topic",
     ["FOREIGN_EXCHANGE_RESERVE", "INTERNATIONAL_INVESTMENT_POSITION", "EXTERNAL_DEBT"],
     "Singularized FX reserves code (#1)."),
    ("INTERNATIONAL", "FOREIGN_INVESTMENT", "topic",
     ["FDI_FLOW", "FDI_STOCK", "PORTFOLIO_FLOW"],
     "FDI in/out -> FDI_FLOW (direction flag). Added FDI_STOCK (position) (#7). PORTFOLIO_FLOWS -> PORTFOLIO_FLOW (#1)."),

    ("DEMOGRAPHICS", "POPULATION_STOCK_AND_STRUCTURE", "topic",
     ["TOTAL_POPULATION", "MEDIAN_AGE", "DEPENDENCY_RATIO", "URBANIZATION_RATE"],
     "Population breakdowns are series flags on TOTAL_POPULATION: age AND group (ethnicity/language/religion). POPULATION_BY_GROUP dropped for consistency with POPULATION_BY_AGE (final-review follow-up)."),
    ("DEMOGRAPHICS", "VITAL_STATISTICS", "topic",
     ["BIRTH_RATE", "DEATH_RATE", "FERTILITY_RATE", "LIFE_EXPECTANCY"],
     "Kept: distinct vital measures (not a direction pair)."),
    ("DEMOGRAPHICS", "MIGRATION", "topic",
     ["MIGRATION_FLOW", "NET_MIGRATION"],
     "Collapsed IMMIGRATION/EMIGRATION -> MIGRATION_FLOW + series flag: direction."),
    ("DEMOGRAPHICS", "HOUSEHOLDS_AND_FAMILIES", "topic",
     ["HOUSEHOLD_COUNT", "AVERAGE_HOUSEHOLD_SIZE", "MARRIAGE_RATE", "DIVORCE_RATE"],
     "MARRIAGE/DIVORCE kept here (nuptiality = household-formation theme); decided (#5)."),

    ("HEALTH", "HEALTH_STATUS_AND_OUTCOMES", "topic",
     ["DISEASE_PREVALENCE", "MORTALITY_RATE"],
     "Collapsed MORTALITY_BY_CAUSE + INFANT_MORTALITY -> MORTALITY_RATE + series flag: breakdown (cause/age)."),
    ("HEALTH", "HEALTH_RISK_FACTORS", "topic",
     ["OBESITY_RATE", "SMOKING_RATE", "ALCOHOL_CONSUMPTION", "UNDERNOURISHMENT_RATE"],
     "Kept distinct factors (could be one PREVALENCE concept + factor flag = open question)."),
    ("HEALTH", "HEALTH_EXPENDITURE", "concept", [],
     "Collapsed to one concept -> series flag: per-capita normalization. (Resolves earlier _TOTAL naming.)"),
    ("HEALTH", "HEALTH_SERVICES", "topic",
     ["HOSPITAL_BED", "PHYSICIAN", "IMMUNIZATION_RATE"],
     "Singular concept codes (final review); per-capita = series flag."),

    ("EDUCATION", "ENROLLMENT_AND_PARTICIPATION", "concept", [],
     "Series flags: education level (primary/secondary/tertiary) + measure (enrolment-ratio vs participation-rate) (final review)."),
    ("EDUCATION", "ATTAINMENT_AND_LITERACY", "topic",
     ["LITERACY_RATE", "TERTIARY_ATTAINMENT", "MEAN_YEARS_SCHOOLING"], ""),
    ("EDUCATION", "EDUCATION_EXPENDITURE", "concept", [],
     "Collapsed to one concept -> series flag: per-student/per-capita normalization. (Resolves earlier _TOTAL naming.)"),
    ("EDUCATION", "EDUCATION_OUTCOMES", "topic",
     ["STUDENT_ASSESSMENT_SCORE", "COMPLETION_RATE"], ""),

    ("SOCIETY", "INCOME_AND_INEQUALITY", "topic",
     ["GINI_COEFFICIENT", "POVERTY_RATE", "MEDIAN_HOUSEHOLD_INCOME", "TOP_INCOME_SHARE"], ""),
    ("SOCIETY", "CRIME_AND_JUSTICE", "topic",
     ["CRIME_RATE", "INCARCERATION_RATE"],
     "Collapsed HOMICIDE_RATE -> series flag: offense type on CRIME_RATE."),
    ("SOCIETY", "SOCIAL_PROTECTION", "topic",
     ["SOCIAL_SPENDING"],
     "Collapsed PENSION_EXPENDITURE -> series flag: function (COFOG) on SOCIAL_SPENDING."),
    ("SOCIETY", "CULTURE_AND_RECREATION", "topic",
     ["CULTURAL_EXPENDITURE", "CULTURAL_PARTICIPATION"], ""),
    ("SOCIETY", "WELLBEING", "concept", [], ""),
    ("SOCIETY", "HOUSING_CONDITIONS", "topic",
     ["HOMEOWNERSHIP_RATE", "OVERCROWDING_RATE", "HOMELESSNESS", "DWELLING_STOCK"], ""),
    ("SOCIETY", "BASIC_SERVICES_ACCESS", "concept", [],
     "Collapsed to one concept -> series flag: service (water/sanitation/electricity)."),

    ("ENVIRONMENT", "EMISSIONS_AND_AIR_QUALITY", "topic",
     ["GHG_EMISSIONS", "AIR_QUALITY"],
     "Collapsed CO2 -> GHG_EMISSIONS (gas flag); generalized PM25 -> AIR_QUALITY (pollutant flag). Two distinct measurement kinds: emission flow vs ambient concentration."),
    ("ENVIRONMENT", "NATURAL_RESOURCES", "topic",
     ["WATER_WITHDRAWAL", "FOREST_AREA", "LAND_USE"], ""),
    ("ENVIRONMENT", "CLIMATE", "topic",
     ["AVERAGE_TEMPERATURE", "PRECIPITATION"], ""),
    # ENVIRONMENTAL_ACCOUNTS dropped (final review): its 3 concepts were normalizations/
    # components homed elsewhere -- ENERGY_INTENSITY -> /GDP flag on ENERGY_CONSUMPTION;
    # RENEWABLE_ENERGY_SHARE -> by-source flag on ENERGY_PRODUCTION; ENVIRONMENTAL_TAX_REVENUE
    # -> tax-type flag on TAX_REVENUE. SEEA "accounts" view = a future series_collection.

    ("OTHER", "ENERGY", "topic",
     ["ENERGY_CONSUMPTION", "ENERGY_PRODUCTION", "ELECTRICITY_GENERATION"],
     "PARKED. Consumption/production kept (distinct balance items)."),
    ("OTHER", "SCIENCE_AND_TECHNOLOGY", "topic",
     ["RESEARCH_AND_DEVELOPMENT_EXPENDITURE", "PATENT_APPLICATION", "INTERNET_PENETRATION"],
     "PARKED. R_AND_D -> RESEARCH_AND_DEVELOPMENT (off abbrev allow-list) (final review). R&D-vs-ICT split deferred (#9)."),
    ("OTHER", "TOURISM", "topic", ["TOURIST_ARRIVAL", "TOURISM_RECEIPT"], "PARKED. Singularized (#1)."),
    ("OTHER", "TRANSPORT", "topic",
     ["FREIGHT_VOLUME", "PASSENGER_TRAFFIC", "VEHICLE_REGISTRATION"], "PARKED. Singularized (#1)."),
    ("OTHER", "UNCLASSIFIED", "concept", [], ""),
]

ROOTS = [
    ("PRICES", "Prices"),
    ("NATIONAL_ACCOUNTS", "National Accounts"),
    ("PRODUCTION_BUSINESS_ACTIVITY", "Production & Business Activity"),
    ("RETAIL_CONSUMPTION", "Retail & Consumption"),
    ("LABOR", "Labour"),
    ("MONETARY_BANKING", "Monetary & Banking"),
    ("FINANCIAL_INDICATORS", "Financial Indicators"),
    ("GOVERNMENT_FISCAL", "Government & Fiscal"),
    ("INTERNATIONAL", "International / External Sector"),
    ("DEMOGRAPHICS", "Demographics"),
    ("HEALTH", "Health"),
    ("EDUCATION", "Education"),
    ("SOCIETY", "Society & Living Conditions"),
    ("ENVIRONMENT", "Environment"),
    ("OTHER", "Other"),
]

CRITIQUE = [
    ("blocker", "FINANCIAL_INDICATORS>BOND_YIELDS_AND_SPREADS", "GOVT_BOND_YIELD_10Y/_2Y mint a concept per maturity; maturity is a variant; under-complete (30Y/3M homeless).", "Collapse to GOVERNMENT_BOND_YIELD with maturity as a series flag. [APPLIED]"),
    ("major", "EDUCATION>EDUCATION_OUTCOMES", "PISA_SCORE is an OECD program label, not a universal function (M2/M3 analogue).", "Rename STUDENT_ASSESSMENT_SCORE; program as series flag. [APPLIED]"),
    ("major", "FINANCIAL/MONETARY", "No home for interbank/money-market reference rates (SOFR/EURIBOR/SONIA).", "Add INTERBANK_RATE. [APPLIED under BANKING_SECTOR]"),
    ("major", "NATIONAL_ACCOUNTS>EXPENDITURE_COMPONENTS", "NET_EXPORTS minted but exports/imports are the SNA components; CHANGES_IN_INVENTORIES missing.", "Add CHANGES_IN_INVENTORIES [APPLIED]; EXPORTS/IMPORTS split [open question]."),
    ("major", "INTERNATIONAL>BALANCE_OF_PAYMENTS", "Remittances (personal transfers) homeless — major EM series.", "Add REMITTANCES. [APPLIED]"),
    ("major", "GOVERNMENT_FISCAL>EXPENDITURE", "Government interest payments / debt service homeless; thin subdomain.", "Add INTEREST_PAYMENTS. [APPLIED]"),
    ("major", "GOVERNMENT_FISCAL/SOCIETY/HEALTH/EDUCATION", "Systematic concept-per-normalization: 8 '_GDP' ratio concepts beside their level.", "Demote all to series unit/flag; keep level only. [APPLIED]"),
    ("major", "SOCIETY>INCOME_AND_INEQUALITY", "INCOME_SHARE_TOP10 bakes a quantile parameter into identity.", "Rename TOP_INCOME_SHARE; quantile as series flag. [APPLIED]"),
    ("minor", "PRICES>CONSUMER_PRICES", "CPI_FOOD/ENERGY/SERVICES are COICOP sub-index breakdowns (slippery slope).", "Justify as watched headlines or fold to series component flag. [OPEN]"),
    ("minor", "PRODUCTION>SECTORAL_OUTPUT", "STEEL_PRODUCTION over-granular among sector aggregates.", "Drop; commodity as series flag. [APPLIED]"),
    ("minor", "PRODUCTION/NATIONAL_ACCOUNTS", "Business inventory STOCK / inventory-to-sales homeless (distinct from the flow).", "Add a stock concept or accept accretion. [OPEN]"),
    ("minor", "INTERNATIONAL>FOREIGN_INVESTMENT", "FDI STOCK missing; inflows/outflows is a direction variant; inconsistent vs migration.", "Decide direction convention tree-wide; add FDI_STOCK. [OPEN]"),
    ("minor", "LABOR/GOVERNMENT_FISCAL", "Government / public-sector / by-sector employment homeless.", "Confirm EMPLOYMENT_LEVEL absorbs via sector flag, or add EMPLOYMENT_BY_SECTOR. [OPEN]"),
    ("minor", "NATIONAL_ACCOUNTS>INPUT_OUTPUT", "I-O tables are a framework, not a scalar concept; risks placeholder.", "Kept as concept-leaf. [DECIDED #8 / final review]"),
    ("minor", "OTHER>SCIENCE_AND_TECHNOLOGY", "INTERNET_PENETRATION (ICT access) bundled with R&D inputs/outputs.", "Split R&D vs ICT when promoted out of OTHER. [OPEN]"),
    ("nit", "Tree-wide (concept codes)", "Pluralization inconsistent (FDI_INFLOWS plural vs EMPLOYMENT_LEVEL singular).", "Adopt one rule (recommend singular) and seed against it. [OPEN]"),
    ("nit", "RESERVES_AND_IIP vs PUBLIC_DEBT", "EXTERNAL_DEBT vs EXTERNAL_PUBLIC_DEBT adjacent / conflatable.", "Confirm split intent; total external debt under IIP. [APPLIED: dropped EXTERNAL_PUBLIC_DEBT]"),
]

QUALITY = [
    ("major", "8 fiscal/social ratio concepts", "Bare '_GDP' suffix reads 'GDP of X'; inconsistent with the '_TO_' connector in CREDIT_TO_PRIVATE_SECTOR.", "Either drop (ratio = unit) or standardize _TO_GDP. [APPLIED: dropped]"),
    ("major", "BOND_YIELDS_AND_SPREADS", "GOVT abbreviated only here; everywhere else GOVERNMENT spelled out.", "GOVERNMENT_BOND_YIELD. [APPLIED]"),
    ("major", "SCIENCE_AND_TECHNOLOGY", "RD initialism inconsistent / ambiguous.", "R_AND_D_EXPENDITURE. [APPLIED]"),
    ("major", "RESERVES_AND_IIP:FX_RESERVES", "FX abbreviated while sibling spells EXCHANGE_RATE.", "FOREIGN_EXCHANGE_RESERVES. [APPLIED]"),
    ("major", "COMMODITY_PRICES:AGRI_PRICE_INDEX", "AGRI abbreviated vs AGRICULTURAL_OUTPUT.", "AGRICULTURAL_PRICE_INDEX. [APPLIED]"),
    ("major", "CONSUMER_PRICES:CPI_CORE", "Core CPI is a methodological variant.", "Quality says drop; OVERRULED by ADR 0026 §5.2 (CPI_CORE is the canonical kept concept)."),
    ("major", "GDP_AND_GROWTH:GDP_GROWTH_RATE", "Growth rate is a transform of GDP_REAL; _RATE vs level inconsistency.", "Drop. [APPLIED]"),
    ("major", "CREDIT_AND_DEBT:MORTGAGE_CREDIT", "_CREDIT noun mixes with _DEBT siblings.", "MORTGAGE_DEBT. [APPLIED]; subdomain-wide harmonization [OPEN]."),
    ("major", "EXTERNAL_DEBT vs EXTERNAL_PUBLIC_DEBT", "Near-duplicate across two roots; double-homes the idea.", "Keep EXTERNAL_DEBT; drop EXTERNAL_PUBLIC_DEBT. [APPLIED]"),
    ("major", "RESERVES_AND_IIP", "L2 abbreviates IIP while child spells it out.", "RESERVES_AND_INVESTMENT_POSITION. [APPLIED]"),
    ("minor", "VOLATILITY_AND_RISK:SOVEREIGN_CDS_SPREAD", "CDS is the only bare initialism among spelled-out neighbours.", "Spell out or sanction on allow-list. [OPEN]"),
    ("minor", "EMISSIONS:CO2/GHG/PM25", "Bare scientific initialisms vs spell-out convention.", "Sanction on allow-list. [OPEN]"),
    ("minor", "EQUITY_MARKETS:STOCK_MARKET_INDEX", "STOCK vs EQUITY lexical inconsistency.", "EQUITY_MARKET_INDEX. [APPLIED]"),
    ("minor", "HOUSEHOLD_CONSUMPTION vs HOUSEHOLD_EXPENDITURE_TOTAL", "Same magnitude from two angles; classifier confusion.", "Disambiguate grain in names or confirm two homes. [OPEN]"),
    ("minor", "EXPENDITURE_BY_CATEGORY / POPULATION_BY_AGE", "_BY_X is a breakdown dimension, not a concept.", "Drop EXPENDITURE_BY_CATEGORY [APPLIED]; reconsider POPULATION_BY_AGE [OPEN]."),
    ("minor", "SECTORAL_OUTPUT / COMMODITY_PRICES", "Per-commodity leaves break sibling aggregation level.", "Decide per-commodity vs per-group rule. [partly APPLIED: dropped STEEL]"),
    ("minor", "INCOME_AND_SAVING saving _RATE", "_RATE bakes a transform into identity.", "Name NATIONAL_SAVING or apply rate/level discipline. [OPEN]"),
    ("minor", "MARRIAGE_RATE/DIVORCE_RATE", "Registered events split from births/deaths across two L2s.", "Move to VITAL_STATISTICS or document. [OPEN]"),
    ("minor", "LABOR_PRODUCTIVITY (L2==concept)", "Identical L2 and concept code -> UNIQUE violation.", "Rename L2 to PRODUCTIVITY. [APPLIED]"),
    ("minor", "PROPERTY_PRICES suffix", "_PRICE vs _PRICE_INDEX within one L2.", "Harmonize to _PRICE_INDEX. [OPEN]"),
    ("minor", "PER_CAPITA normalization", "Per-capita is a concept in some places, absent in siblings.", "Decide concept vs series transform uniformly. [OPEN]"),
    ("minor", "RETAIL_SALES_VALUE/VOLUME", "Nominal/real basis as two concepts.", "Consider single concept + basis flag. [OPEN]"),
    ("nit", "GDP_NOMINAL/GDP_REAL", "Price basis in code; defensible but inconsistent with treating real/nominal as flag.", "Document the GDP-only exception. [OPEN]"),
]

# Refreshed after the final review (wwfxszcgz) -- most earlier items resolved by the
# collapse / singular / final-review passes. Remaining genuinely open or deferred:
OPEN_QUESTIONS = [
    "DEFERRED: when SCIENCE_AND_TECHNOLOGY is promoted out of OTHER, split R&D/innovation from ICT/digital-access.",
    "DOWNSTREAM (series layer): formally define the series-flag vocabulary the tree now leans on -- sector, direction (in/out/net), basis (nominal/real, value/volume), normalization (per-capita, per-area, /GDP), component/breakdown, substance/commodity/gas/pollutant, maturity, quantile, measure. The onboarding router needs this to place series deterministically.",
    "AT SEED TIME (descriptions): write concept descriptions that encode grain where names alone are ambiguous -- esp. HOUSEHOLD_SPENDING (survey) vs HOUSEHOLD_CONSUMPTION (SNA), and ENROLLMENT_AND_PARTICIPATION's measure flag.",
    "FUTURE (series_collections): cross-cutting reassembly views that the strict tree intentionally does not hold -- housing, and the SEEA environmental-accounts view (after dropping ENVIRONMENTAL_ACCOUNTS).",
]

ACRONYMS = {"CPI", "PPI", "GDP", "GNI", "PMI", "FDI", "CDS", "CO2", "GHG",
            "PM25", "IIP", "NPL", "ULC", "REER", "NEER", "ICT"}

# Concept codes are singular (the key); display names keep the idiomatic plural.
NAME_OVERRIDES = {
    "HOUSING_START": "Housing Starts",
    "BUILDING_PERMIT": "Building Permits",
    "BUSINESS_REGISTRATION": "Business Registrations",
    "BUSINESS_BANKRUPTCY": "Business Bankruptcies",
    "INTEREST_PAYMENT": "Interest Payments",
    "PATENT_APPLICATION": "Patent Applications",
    "MOBILE_MONEY_ACCOUNT": "Mobile Money Accounts",
    "TOURIST_ARRIVAL": "Tourist Arrivals",
    "TOURISM_RECEIPT": "Tourism Receipts",
    "VEHICLE_REGISTRATION": "Vehicle Registrations",
    "BANK_RESERVE": "Bank Reserves",
    "FOREIGN_EXCHANGE_RESERVE": "Foreign Exchange Reserves",
    "BANK_DEPOSIT": "Bank Deposits",
    "CENTRAL_BANK_ASSET": "Central Bank Assets",
    "NON_PERFORMING_LOAN": "Non-Performing Loans",
    "CORPORATE_PROFIT": "Corporate Profits",
    "CHANGE_IN_INVENTORIES": "Changes in Inventories",
    "PORTFOLIO_FLOW": "Portfolio Flows",
    "FDI_FLOW": "FDI Flows",
    "INVENTORIES": "Inventories",
    "PHYSICIAN": "Physicians",
    "HOSPITAL_BED": "Hospital Beds",
    "FINANCIAL_ACCESS_POINT": "Financial Access Points",
    "COMMODITY_PRICE": "Commodity Prices",
    "POLICY_RATE": "Policy Rate",
    "EXCHANGE_RATE": "Exchange Rate",
}


def title(code: str) -> str:
    parts = code.split("_")
    out = []
    for p in parts:
        if p in ACRONYMS:
            out.append(p)
        elif p == "AND":
            out.append("&")
        else:
            out.append(p.capitalize())
    return " ".join(out)


def disp(code: str) -> str:
    return NAME_OVERRIDES.get(code, title(code))


# ---- workbook ----
wb = Workbook()

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(bold=True, color="FFFFFF")
L1_FILL = PatternFill("solid", fgColor="D6DCE5")
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
SEV_FILL = {
    "blocker": PatternFill("solid", fgColor="C00000"),
    "major": PatternFill("solid", fgColor="ED7D31"),
    "minor": PatternFill("solid", fgColor="FFE699"),
    "nit": PatternFill("solid", fgColor="E2EFDA"),
}


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = WRAP
        cell.border = BORDER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# Sheet 1: Taxonomy grid (L1/L2/L3)
ws = wb.active
ws.title = "Taxonomy (L1-L2-L3)"
ws.append(["L1 domain", "L2 subdomain", "L2 kind", "L3 concepts (suggested)", "# L3", "Notes"])
last_l1 = None
for l1, l2, kind, l3, note in TAXONOMY:
    show_l1 = l1 if l1 != last_l1 else ""
    last_l1 = l1
    ws.append([show_l1, l2, kind, ", ".join(l3) if l3 else "(concept-leaf)", len(l3), note])
    r = ws.max_row
    for c in range(1, 7):
        ws.cell(row=r, column=c).alignment = WRAP
        ws.cell(row=r, column=c).border = BORDER
    if show_l1:
        ws.cell(row=r, column=1).fill = L1_FILL
        ws.cell(row=r, column=1).font = Font(bold=True)
    if kind == "concept":
        ws.cell(row=r, column=3).font = Font(italic=True, color="7030A0")
style_header(ws, 6)
set_widths(ws, [30, 34, 10, 60, 6, 50])

# Sheet 2: Flat seed rows (code, name, kind, parent_code, level)
ws2 = wb.create_sheet("Seed rows (flat)")
ws2.append(["code", "name", "kind", "parent_code", "level"])
seen = set()
dups = []
for code, name in ROOTS:
    ws2.append([code, name, "topic", "", 1])
    seen.add(code)
for l1, l2, kind, l3, note in TAXONOMY:
    if l2 in seen:
        dups.append(l2)
    seen.add(l2)
    ws2.append([l2, disp(l2), kind, l1, 2])
    for c in l3:
        if c in seen:
            dups.append(c)
        seen.add(c)
        ws2.append([c, disp(c), "concept", l2, 3])
for r in range(2, ws2.max_row + 1):
    for c in range(1, 6):
        ws2.cell(row=r, column=c).alignment = TOP
        ws2.cell(row=r, column=c).border = BORDER
    lvl = ws2.cell(row=r, column=5).value
    if lvl == 1:
        for c in range(1, 6):
            ws2.cell(row=r, column=c).fill = L1_FILL
        ws2.cell(row=r, column=1).font = Font(bold=True)
style_header(ws2, 5)
set_widths(ws2, [38, 38, 10, 32, 7])

# Sheet 3: Review findings (critique + quality)
ws3 = wb.create_sheet("Review findings")
ws3.append(["source", "severity", "location", "issue", "recommendation / status"])
for src, rows in (("design-critic", CRITIQUE), ("dedup-quality", QUALITY)):
    for sev, loc, issue, rec in rows:
        ws3.append([src, sev, loc, issue, rec])
        r = ws3.max_row
        for c in range(1, 6):
            ws3.cell(row=r, column=c).alignment = WRAP
            ws3.cell(row=r, column=c).border = BORDER
        ws3.cell(row=r, column=2).fill = SEV_FILL.get(sev, PatternFill())
        if sev in ("blocker", "major"):
            ws3.cell(row=r, column=2).font = Font(bold=True, color="FFFFFF" if sev != "minor" else "000000")
style_header(ws3, 5)
set_widths(ws3, [15, 10, 38, 62, 62])

# Sheet 4: Open questions
ws4 = wb.create_sheet("Open questions")
ws4.append(["#", "Decision still owed to the operator"])
for i, q in enumerate(OPEN_QUESTIONS, 1):
    ws4.append([i, q])
    r = ws4.max_row
    for c in range(1, 3):
        ws4.cell(row=r, column=c).alignment = WRAP
        ws4.cell(row=r, column=c).border = BORDER
style_header(ws4, 2)
set_widths(ws4, [5, 110])

# Sheet 5: README
ws5 = wb.create_sheet("README", 0)
lines = [
    ("macrodb category taxonomy", True),
    ("Source: ADR 0026; review (w0ul9g3el) + coverage validation (wz3k53mcb), 2026-06-18.", False),
    ("Coverage: 94% of 388 real series across 12 sources mapped cleanly (see ADR 0026 Validation).", False),
    ("", False),
    ("Sheets:", True),
    ("- Taxonomy (L1-L2-L3): the browse grid. L2 'concept' kind = attachable concept-leaf (no L3).", False),
    ("- Seed rows (flat): (code, name, kind, parent_code, level) — the seed/data scaffold shape.", False),
    ("- Review findings: design-critic + dedup/quality issues, with [APPLIED]/[OPEN]/[OVERRULED] status.", False),
    ("- Open questions: debatable calls NOT auto-resolved; for the operator to decide.", False),
    ("", False),
    ("Rules (ADR 0026):", True),
    ("- Strict single-parent tree, depth <= 3. Codes UPPERCASE SCREAMING_SNAKE, unique.", False),
    ("- Concepts name the economic FUNCTION, not a country label (BROAD_MONEY not M2/M3/M4).", False),
    ("- Methodological variants are SERIES-level flags, never a concept each.", False),
    ("- 'name' column is a generated first-pass display label; refine freely.", False),
    ("- L3 concepts are SUGGESTED universal seeds; the long tail accretes via onboarding.", False),
    ("", False),
    ("Counts:", True),
]
roots_n = len(ROOTS)
l2_n = len(TAXONOMY)
l3_n = sum(len(t[3]) for t in TAXONOMY)
lines.append((f"- {roots_n} roots, {l2_n} L2 buckets, {l3_n} suggested L3 concepts.", False))
for text, bold in lines:
    ws5.append([text])
    if bold:
        ws5.cell(row=ws5.max_row, column=1).font = Font(bold=True)
ws5.column_dimensions["A"].width = 110

import os
out = os.path.join(os.path.dirname(__file__), "category_taxonomy.xlsx")
wb.save(out)
print("wrote", out)
print("roots", roots_n, "L2", l2_n, "L3", l3_n, "total nodes", roots_n + l2_n + l3_n)
print("duplicate codes:", sorted(set(dups)) or "NONE")
